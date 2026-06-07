from __future__ import annotations

import io
import re
from datetime import date

import pandas as pd
from openpyxl import load_workbook

from services.executive_report_exports import (
    MARGIN_INCHES,
    executive_report_to_excel_bytes,
    executive_report_to_pdf_bytes,
)


def _sections() -> dict[str, pd.DataFrame]:
    sample = pd.DataFrame(
        [
            {"Sucursal": "CENTRAL", "VentaNetaQ": 1000.0, "%Margen": 0.60},
            {"Sucursal": "NORTE", "VentaNetaQ": 500.0, "%Margen": 0.53},
        ]
    )
    return {
        "R-DASH-01": pd.DataFrame([{"Indicador": "Venta Neta Q", "Valor": 1500.0}]),
        "T-DASH-01": sample,
        "T-GER-05": sample,
        "T-GER-04": sample,
        "T-GER-03": sample,
        "T-EXI-02": sample,
    }


def test_executive_excel_has_expected_sheets_and_print_settings() -> None:
    content = executive_report_to_excel_bytes(
        _sections(),
        date(2026, 6, 1),
        date(2026, 6, 7),
    )
    workbook = load_workbook(io.BytesIO(content), read_only=False)

    assert workbook.sheetnames == list(_sections())
    worksheet = workbook["T-DASH-01"]
    assert worksheet.page_setup.orientation == "landscape"
    assert worksheet.page_setup.paperSize == 1
    assert abs(worksheet.page_margins.left - MARGIN_INCHES) < 0.001
    assert abs(worksheet.page_margins.right - MARGIN_INCHES) < 0.001
    assert abs(worksheet.page_margins.top - MARGIN_INCHES) < 0.001
    assert abs(worksheet.page_margins.bottom - MARGIN_INCHES) < 0.001
    assert "Codigo: T-DASH-01" in worksheet["A2"].value


def test_executive_pdf_is_letter_landscape_and_contains_report_codes() -> None:
    content = executive_report_to_pdf_bytes(
        _sections(),
        date(2026, 6, 1),
        date(2026, 6, 7),
    )
    assert content.startswith(b"%PDF-")
    assert re.search(rb"/MediaBox\s*\[\s*0\s+0\s+792\s+612\s*\]", content)
    assert len(re.findall(rb"/Type\s*/Page(?!s)", content)) == 6
